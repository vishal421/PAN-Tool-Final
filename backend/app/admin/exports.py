"""Generic tabular export helpers (CSV + Excel) shared by every admin export endpoint."""

from __future__ import annotations

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2A2F3D", end_color="2A2F3D", fill_type="solid")
_BODY_FONT = Font(name="Arial", size=10)


def rows_to_csv(rows: list[dict], columns: list[str] | None = None) -> str:
    columns = columns or (list(rows[0].keys()) if rows else [])
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({c: row.get(c, "") for c in columns})
    return buf.getvalue()


def rows_to_xlsx(rows: list[dict], columns: list[str] | None = None, sheet_title: str = "Export") -> bytes:
    columns = columns or (list(rows[0].keys()) if rows else [])
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet name length limit

    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(columns, start=1):
            value = row.get(header, "")
            if hasattr(value, "isoformat"):
                value = value.isoformat(sep=" ", timespec="seconds")
            ws.cell(row=row_idx, column=col_idx, value=value).font = _BODY_FONT

    for col_idx, header in enumerate(columns, start=1):
        max_len = max([len(str(header))] + [len(str(r.get(header, ""))) for r in rows]) if rows else len(header)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 50)
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
