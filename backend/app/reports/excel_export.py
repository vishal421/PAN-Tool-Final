"""
Excel export for the configuration summary.

Plain tabular data, no formulas - each sheet is a straight dump of one
object type's rows (from app/reports/summary.py) so it matches the JSON
summary the UI table renders exactly. Kept simple deliberately: this is a
migration inventory report, not a financial model, so there's nothing to
recalculate.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.normalizer.models import NormalizedConfig
from app.reports.summary import build_summary

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2A2F3D", end_color="2A2F3D", fill_type="solid")
_BODY_FONT = Font(name="Arial", size=10)
_TITLE_FONT = Font(name="Arial", bold=True, size=14)

_SHEET_TITLES = {
    "addresses": "Addresses",
    "address_groups": "Address Groups",
    "services": "Services",
    "service_groups": "Service Groups",
    "interfaces": "Interfaces",
    "routes": "Routes",
    "policies": "Security Policies",
    "nat_rules": "NAT Rules",
}


def _write_table(ws, rows: list[dict], start_row: int = 1) -> None:
    if not rows:
        ws.cell(row=start_row, column=1, value="No objects of this type were found in the source config.").font = _BODY_FONT
        return

    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _BODY_FONT

    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(r.get(header, ""))) for r in rows])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 60)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def build_summary_workbook(config: NormalizedConfig, vendor: str, source_filename: str) -> Workbook:
    summary = build_summary(config)
    wb = Workbook()

    overview = wb.active
    overview.title = "Overview"
    overview["A1"] = "Firewall Configuration Summary"
    overview["A1"].font = _TITLE_FONT
    overview["A2"] = f"Source vendor: {vendor}"
    overview["A2"].font = _BODY_FONT
    overview["A3"] = f"Source file: {source_filename}"
    overview["A3"].font = _BODY_FONT

    overview.cell(row=5, column=1, value="Object Type").font = _HEADER_FONT
    overview.cell(row=5, column=1).fill = _HEADER_FILL
    overview.cell(row=5, column=2, value="Count").font = _HEADER_FONT
    overview.cell(row=5, column=2).fill = _HEADER_FILL

    count_labels = [
        ("addresses", "Addresses"),
        ("address_groups", "Address Groups"),
        ("services", "Services"),
        ("service_groups", "Service Groups"),
        ("interfaces", "Interfaces"),
        ("routes", "Routes"),
        ("policies", "Security Policies"),
        ("nat_rules", "NAT Rules"),
        ("warnings", "Warnings"),
        ("errors", "Errors"),
        ("unsupported", "Unsupported Items"),
    ]
    for i, (key, label) in enumerate(count_labels, start=6):
        overview.cell(row=i, column=1, value=label).font = _BODY_FONT
        overview.cell(row=i, column=2, value=summary["counts"].get(key, 0)).font = _BODY_FONT

    overview.column_dimensions["A"].width = 22
    overview.column_dimensions["B"].width = 12

    for key, title in _SHEET_TITLES.items():
        ws = wb.create_sheet(title=title)
        _write_table(ws, summary["tables"].get(key, []))

    return wb
